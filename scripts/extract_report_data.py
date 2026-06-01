from __future__ import annotations

import io
import json
import math
import os
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
MARKETING_ID = "1IuErSQZMfRLlPBhiG1uudhggKqrrXDOyffMGn4QV7GI"
FINANCIAL_ID = "1W8q8cNgpr99vyj_VCSWl7JpOWYs__NY1vipRup35RZ8"
LOCAL_SOURCE = ROOT / "data-source"
OUT = ROOT / "public" / "report" / "data" / "report-data.json"
OUT_JS = ROOT / "public" / "report" / "data" / "report-data.js"

MONTHS = {
    "january",
    "february",
    "march",
    "april",
    "may",
    "june",
    "july",
    "august",
    "september",
    "october",
    "november",
    "december",
}


def clean(value):
    if value is None:
        return None
    if isinstance(value, float) and (math.isnan(value) or math.isinf(value)):
        return None
    if isinstance(value, str):
        stripped = value.strip()
        return stripped or None
    return value


def row_values(row):
    return [clean(cell.value) for cell in row]


def normalize_header(header):
    labels = []
    seen = {}
    for index, value in enumerate(header):
        label = str(value).strip() if value not in (None, "") else f"column_{index}"
        label = label.replace("\n", " ")
        count = seen.get(label, 0)
        seen[label] = count + 1
        labels.append(label if count == 0 else f"{label}_{count + 1}")
    return labels


def record_from(labels, row):
    return {
        label: value
        for label, value in zip(labels, row)
        if value is not None and not str(label).startswith("column_")
    }


def parse_sheet(ws):
    rows = [row_values(row) for row in ws.iter_rows()]
    rows = [row for row in rows if any(value is not None for value in row)]
    if not rows:
        return {"headers": [], "monthly": [], "totals": None, "otherRows": []}

    headers = normalize_header(rows[0])
    monthly = []
    totals = None
    other = []

    for raw in rows[1:]:
        first = clean(raw[0])
        if first is None:
            continue
        first_text = str(first).strip()
        record = record_from(headers, raw)
        if first_text.lower() in MONTHS:
            record["Month"] = first_text
            monthly.append(record)
        elif first_text.upper().startswith("TOTAL"):
            totals = record
        else:
            other.append(record)

    return {
        "headers": headers,
        "monthly": monthly,
        "totals": totals,
        "otherRows": other,
    }


def parse_workbook_bytes(content, file_name, workbook_type):
    wb = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    parsed = {
        "file": file_name,
        "type": workbook_type,
        "sheets": {},
        "sheetNames": list(wb.sheetnames),
    }
    for ws in wb.worksheets:
        parsed["sheets"][ws.title] = parse_sheet(ws)
    return parsed


def download_sheet_xlsx(file_id):
    credentials_json = os.environ.get("GOOGLE_SERVICE_ACCOUNT_JSON")
    if not credentials_json:
        raise RuntimeError("GOOGLE_SERVICE_ACCOUNT_JSON is required when data-source XLSX files are absent.")

    from google.oauth2 import service_account
    from googleapiclient.discovery import build

    credentials = service_account.Credentials.from_service_account_info(
        json.loads(credentials_json),
        scopes=["https://www.googleapis.com/auth/drive.readonly"],
    )
    drive = build("drive", "v3", credentials=credentials)
    request = drive.files().export_media(
        fileId=file_id,
        mimeType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    return request.execute()


def workbook_source(file_name, file_id):
    local = LOCAL_SOURCE / file_name
    if local.exists():
        return local.read_bytes()
    return download_sheet_xlsx(file_id)


def main():
    marketing_bytes = workbook_source("marketing-revops.xlsx", MARKETING_ID)
    financial_bytes = workbook_source("financial-reports.xlsx", FINANCIAL_ID)
    data = {
        "appName": "AW3 Marketing Reporting",
        "generatedAt": datetime.now(timezone.utc).isoformat().replace("+00:00", "Z"),
        "generatedFrom": {
            "marketingWorkbook": "AW3 Marketing & RevOps/Sales Metrics (by FY)",
            "financialWorkbook": "AW3 Financial Reports by FY",
        },
        "marketing": parse_workbook_bytes(marketing_bytes, "marketing-revops.xlsx", "marketing_revops"),
        "financial": parse_workbook_bytes(financial_bytes, "financial-reports.xlsx", "financial_reports"),
    }

    OUT.parent.mkdir(parents=True, exist_ok=True)
    json_text = json.dumps(data, indent=2, ensure_ascii=False)
    OUT.write_text(json_text, encoding="utf-8")
    OUT_JS.write_text(f"window.REPORT_DATA = {json_text};\n", encoding="utf-8")

    marketing_rows = sum(len(sheet["monthly"]) for sheet in data["marketing"]["sheets"].values())
    financial_rows = sum(len(sheet["monthly"]) for sheet in data["financial"]["sheets"].values())
    print(json.dumps({
        "output": str(OUT),
        "marketingMonthlyRows": marketing_rows,
        "financialMonthlyRows": financial_rows,
    }, indent=2))


if __name__ == "__main__":
    main()
